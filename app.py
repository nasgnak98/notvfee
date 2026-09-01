import io
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pandas as pd
import streamlit as st

# --- 페이지 기본 설정 ---
st.set_page_config(
    layout="wide",
    page_title="미부과세대 종합 시스템",
    page_icon="🏢",
)


# --- 외부 CSS 로드 함수 ---
def load_css(file_name: str = "style.css"):
  """CSS 파일을 읽어와 Streamlit 화면에 적용합니다."""
  css_path = Path(__file__).parent / file_name
  if css_path.exists():
    with open(css_path, "r", encoding="utf-8") as f:
      st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)


load_css("style.css")

# --- 엑셀 스타일 상수 선언 ---
FONT_TITLE = Font(name="돋움", size=11, bold=True)
FONT_HEADER = Font(name="돋움", size=11, bold=True)
FONT_BODY = Font(name="맑은 고딕", size=10)

FILL_HEADER = PatternFill(
    start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
)
FILL_HIGHLIGHT = PatternFill(
    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
)

THIN_SIDE = Side(style="thin", color="000000")
BORDER_ALL = Border(
    left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")

DEFAULT_NOTICE_ROWS = [
    "◆ 담당 이강산 연락처 ☎ 740-7143 FAX 740-7159 Email kbsjejutv@kbs.co.kr",
    "◆ 아래 내용 2번 항목은 지난달 부과대수 확정이후 변동이 있는 세대 내역이며,",
    "   3번 항목은 2번 항목 변동사항의 미부과 세대를 포함하고, 부과세대를 제외한",
    "   이번달 수신료 미부과세대 내역입니다. 전기면제와 공가세대를 추가해",
    "   제출해주시면 특별한 사유가 없는 한 주신 자료로 확정하겠습니다.",
]

NOTE_MAP = {
    "기초수급": "기초수급자",
    "기초생활": "기초수급자",
    "수급자": "기초수급자",
    "시청각": "시청각장애",
    "장애인": "시청각장애",
    "장애": "시청각장애",
    "국가유공": "국가유공자",
    "유공자": "국가유공자",
}


# --- 데이터 처리 및 로직 함수 ---
@st.cache_data(show_spinner=False)
def load_table(file_bytes: bytes) -> pd.DataFrame:
  """엑셀 파일 바이트를 받아 기준 헤더(세대번호) 위치를 탐색 후 DataFrame으로 변환합니다."""
  df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None)
  skip_row = None

  row_strings = df_raw.fillna("").astype(str).agg(" ".join, axis=1)

  for idx, row_str in enumerate(row_strings):
    if "3. 세대 미부과 명세" in row_str or "3. 미부과 세대 명세" in row_str:
      skip_row = idx + 1
      break

  if skip_row is None:
    for idx, row_str in enumerate(row_strings):
      if "주소" in row_str:
        for offset in (1, 2, 3):
          if idx + offset < len(row_strings):
            sub_str = row_strings.iloc[idx + offset]
            if "세대번호" in sub_str or "호수" in sub_str:
              skip_row = idx + offset
              break
        if skip_row is not None:
          break

  if skip_row is None:
    for idx, row_str in enumerate(row_strings):
      if "세대번호" in row_str:
        skip_row = idx
        break

  if skip_row is None:
    raise ValueError(
        "⚠️ 파일에서 기준 헤더(세대번호)를 찾을 수 없습니다. 파일 양식을 확인해"
        " 주세요."
    )

  header_row = df_raw.iloc[skip_row]
  df = df_raw.iloc[skip_row + 1 :].copy()
  df.columns = [str(val).strip() for val in header_row.values]

  key_col = next(
      (
          c
          for c in df.columns
          if any(k in c for k in ["세대번호", "세대", "호수", "동호수"])
      ),
      df.columns[0],
  )
  df["세대번호"] = df[key_col].astype(str).str.strip()

  exclude_vals = {"세대번호", "호수", "동호수", "nan", "None", ""}
  df = df[~df["세대번호"].isin(exclude_vals)]

  return df


def extract_info(row: pd.Series, target_df: pd.DataFrame):
  """행 데이터로부터 변동일자 및 원인 사유를 추출합니다."""
  res_date = "-"
  res_content = "-"

  date_cols = [
      c
      for c in target_df.columns
      if any(
          k in str(c)
          for k in ["말소일", "면제일", "변동일", "등록일", "일자", "일시"]
      )
  ]
  for c in date_cols:
    val = row[c]
    if pd.notna(val) and str(val).strip() not in ["", "nan", "None"]:
      if isinstance(val, pd.Timestamp):
        res_date = val.strftime("%Y-%m-%d")
      else:
        res_date = str(val).split()[0].replace("/", "-")
      break

  reason_cols = [
      c
      for c in target_df.columns
      if any(
          k in str(c)
          for k in [
              "말소사항",
              "면제사항",
              "사유",
              "내용",
              "내역",
              "구분",
              "비고",
          ]
      )
  ]
  for c in reason_cols:
    val = row[c]
    if pd.notna(val) and str(val).strip() not in ["", "nan", "None"]:
      res_content = str(val).strip()
      break

  return res_date, res_content


def determine_report_text_and_note(status: str, raw_reason: str):
  """사유에 따라 보고서 표기 문구 및 비고 키워드를 결정합니다."""
  raw_reason_str = str(raw_reason)
  detected_note = ""

  for kw, official in NOTE_MAP.items():
    if kw in raw_reason_str:
      detected_note = official
      break

  is_exemption = bool(detected_note)

  if status == "신규":
    content = "면제등록" if is_exemption else "미소지 등록"
  elif status == "삭제":
    content = "면제해제" if is_exemption else "미소지해제"
  else:
    content = "-"

  return content, detected_note


@st.cache_data(show_spinner=False)
def generate_full_report(
    template_bytes: bytes, current_bytes: bytes, report_rows: list
) -> bytes:
  """Excel 템플릿과 변동사항 데이터를 결합하여 최종 보고서를 생성합니다."""
  wb_tmpl = openpyxl.load_workbook(io.BytesIO(template_bytes))
  ws_tmpl = wb_tmpl.active

  wb_curr = openpyxl.load_workbook(io.BytesIO(current_bytes))
  ws = wb_curr.active

  tmpl_notice_rows = []
  for r in range(1, ws_tmpl.max_row + 1):
    val = str(ws_tmpl.cell(row=r, column=1).value or "").strip()
    if any(
        k in val
        for k in [
            "2. 변동사항",
            "2.변동사항",
            "3. 세대",
            "3.세대",
            "3. 미부과",
            "번호",
            "세대번호",
            "변동일",
        ]
    ):
      break
    if val and not val.startswith("1."):
      if not val.startswith("◆"):
        val = "   " + val.lstrip()
      tmpl_notice_rows.append(val)

  notice_list = tmpl_notice_rows if tmpl_notice_rows else DEFAULT_NOTICE_ROWS

  for r in range(1, min(10, ws.max_row + 1)):
    for c in range(1, ws.max_column + 1):
      cell_val = str(ws.cell(row=r, column=c).value or "")
      if "KBS" in cell_val and not cell_val.endswith("    "):
        ws.cell(row=r, column=c).value = cell_val + "    "

  address_row_idx = None
  for r in range(1, ws.max_row + 1):
    row_str = " ".join([
        str(ws.cell(row=r, column=c).value or "")
        for c in range(1, ws.max_column + 1)
    ])
    if "주소" in row_str:
      address_row_idx = r
      break

  if address_row_idx is None:
    address_row_idx = 2

  ws.delete_rows(4, amount=1)

  if address_row_idx > 4:
    address_row_idx -= 1

  insert_pos = address_row_idx + 2

  ws.insert_rows(insert_pos)
  insert_pos += 1

  ws.insert_rows(insert_pos)
  cell = ws.cell(row=insert_pos, column=1, value="1. 안내사항")
  cell.font = FONT_TITLE
  insert_pos += 1

  for txt in notice_list:
    ws.insert_rows(insert_pos)
    cell = ws.cell(row=insert_pos, column=1, value=txt)
    cell.font = FONT_BODY
    cell.alignment = ALIGN_LEFT
    insert_pos += 1

  ws.insert_rows(insert_pos)
  insert_pos += 1

  ws.insert_rows(insert_pos)
  cell = ws.cell(row=insert_pos, column=1, value="2. 변동사항")
  cell.font = FONT_TITLE
  insert_pos += 1

  target_h_keys = set()

  if report_rows:
    headers = ["번호", "세대번호", "변동일", "변동내용", "부과여부", "비고"]
    ws.insert_rows(insert_pos)
    for c_idx, h_text in enumerate(headers, 1):
      c = ws.cell(row=insert_pos, column=c_idx, value=h_text)
      c.font = FONT_HEADER
      c.fill = FILL_HEADER
      c.border = BORDER_ALL
      c.alignment = ALIGN_CENTER
    insert_pos += 1

    for idx, r_data in enumerate(report_rows, 1):
      ws.insert_rows(insert_pos)
      row_values = [idx] + r_data[1:]
      for c_idx, val in enumerate(row_values, 1):
        c = ws.cell(
            row=insert_pos, column=c_idx, value=val if val is not None else ""
        )
        c.font = FONT_BODY
        c.border = BORDER_ALL
        c.alignment = ALIGN_CENTER

      if r_data[1]:
        target_h_keys.add(str(r_data[1]).strip())

      insert_pos += 1
  else:
    ws.insert_rows(insert_pos)
    cell = ws.cell(row=insert_pos, column=1, value="◆ 변동 내역이 없습니다.")
    cell.font = FONT_BODY
    cell.alignment = ALIGN_LEFT
    insert_pos += 1

  ws.insert_rows(insert_pos)
  insert_pos += 1

  ws.insert_rows(insert_pos)
  cell = ws.cell(row=insert_pos, column=1, value="3. 세대 미부과 명세")
  cell.font = FONT_TITLE
  spec_start_row = insert_pos

  if target_h_keys:
    for r in range(spec_start_row + 1, ws.max_row + 1):
      row_vals = [
          str(ws.cell(row=r, column=c).value or "").strip()
          for c in range(1, ws.max_column + 1)
      ]
      if any(k in row_vals for k in target_h_keys):
        for c in range(1, ws.max_column + 1):
          cell = ws.cell(row=r, column=c)
          if cell.value is not None or cell.border.left.style is not None:
            cell.fill = FILL_HIGHLIGHT

  output = io.BytesIO()
  wb_curr.save(output)
  return output.getvalue()


# --- 메인 화면 UI ---
st.title("🏢 미부과세대 분석 및 보고서")

st.subheader("📁 필수 파일 업로드 (3개 파일 필수)")
col1, col2, col3 = st.columns(3)

with col1:
  tmpl_file = st.file_uploader(
      "1️⃣ 안내사항 템플릿 (필수)", type=["xlsx", "xls"]
  )
with col2:
  file_a = st.file_uploader("2️⃣ 과거 파일 (필수)", type=["xlsx", "xls"])
with col3:
  file_b = st.file_uploader("3️⃣ 현재 파일 (필수)", type=["xlsx", "xls"])

st.markdown("---")

if tmpl_file and file_a and file_b:
  user_filename = st.text_input(
      "📝 저장할 파일명을 입력해 주세요 (확장자 제외 가능):",
      value="미부과세대_변동_보고서_완성본",
  )

  if st.button("⚡ 분석 및 보고서 생성", type="primary"):
    try:
      with st.spinner("데이터 분석 및 보고서 결합 중..."):
        # 캐싱된 함수 호출 (바이트 데이터 입력)
        df_a = load_table(file_a.getvalue())
        df_b = load_table(file_b.getvalue())

        idx_a = set(df_a["세대번호"].unique())
        idx_b = set(df_b["세대번호"].unique())

        deleted = sorted(list(idx_a - idx_b))
        added = sorted(list(idx_b - idx_a))

        dash_rows = []
        report_rows = []

        for k in deleted:
          sub_df = df_a[df_a["세대번호"] == k]
          if not sub_df.empty:
            row_data = sub_df.iloc[0]
            dt, raw_reason = extract_info(row_data, df_a)

            dash_rows.append({
                "구분": "🔴 삭제",
                "세대번호": k,
                "변동일": dt,
                "내용": raw_reason if raw_reason != "-" else "부과 재개",
                "부과여부": "부과",
            })
            report_content, report_note = determine_report_text_and_note(
                "삭제", raw_reason
            )
            report_rows.append(
                [None, k, dt, report_content, "부과", report_note]
            )

        for k in added:
          sub_df = df_b[df_b["세대번호"] == k]
          if not sub_df.empty:
            row_data = sub_df.iloc[0]
            dt, raw_reason = extract_info(row_data, df_b)

            dash_rows.append({
                "구분": "🟢 신규",
                "세대번호": k,
                "변동일": dt,
                "내용": raw_reason if raw_reason != "-" else "미부과 등록",
                "부과여부": "미부과",
            })
            report_content, report_note = determine_report_text_and_note(
                "신규", raw_reason
            )
            report_rows.append(
                [None, k, dt, report_content, "미부과", report_note]
            )

        excel_bytes = generate_full_report(
            tmpl_file.getvalue(), file_b.getvalue(), report_rows
        )

      st.subheader("📊 웹 대시보드 (변동사항 내역)")
      if dash_rows:
        df_dash = pd.DataFrame(dash_rows)

        st.dataframe(
            df_dash.style.set_properties(**{"text-align": "center"}),
            use_container_width=True,
            hide_index=True,
        )

        tsv_data = df_dash.to_csv(index=False, sep="\t")

        # HTML Component (클립보드 복사 버튼)
        st.components.v1.html(
            f"""
                <style>
                    .copy-btn {{
                        background-color: #008CBA;
                        color: white;
                        padding: 8px 16px;
                        border: none;
                        border-radius: 4px;
                        cursor: pointer;
                        font-weight: bold;
                        font-size: 14px;
                    }}
                    .copy-btn:hover {{
                        background-color: #005f73;
                    }}
                </style>
                <button id="copy-btn" class="copy-btn">📋 대시보드 데이터 클립보드에 복사</button>
                <script>
                document.getElementById('copy-btn').addEventListener('click', function() {{
                    const textToCopy = {repr(tsv_data)};
                    navigator.clipboard.writeText(textToCopy).then(function() {{
                        alert('클립보드에 복사되었습니다! 엑셀 등에 Ctrl+V 로 붙여넣으세요.');
                    }}, function(err) {{
                        alert('복사 실패: ' + err);
                    }});
                }});
                </script>
                """,
            height=50,
        )

      else:
        st.success("🎉 변동 사항이 없습니다.")

      st.divider()
      st.subheader("📥 최종 결과물 다운로드")
      st.success(
          "🎉 보고서 생성이 완료되었습니다! 아래 다운로드 버튼을 클릭해 주세요."
      )

      clean_filename = user_filename.strip()
      if not clean_filename.endswith(".xlsx"):
        clean_filename += ".xlsx"

      st.download_button(
          label=f"📥 '{clean_filename}' 다운로드",
          data=excel_bytes,
          file_name=clean_filename,
          mime=(
              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
          ),
          type="primary",
      )

    except Exception as e:
      st.error(f"⚠️ 실행 중 오류가 발생했습니다: {e}")
else:
  st.info(
      "🚨 3개 파일(1️⃣ 안내사항 템플릿, 2️⃣ 과거 파일, 3️⃣ 현재 파일)을 모두"
      " 업로드해 주세요."
  )
